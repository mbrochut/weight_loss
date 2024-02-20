from datetime import datetime
import re
import pandas as pd
from openpyxl import load_workbook
from dateutil.parser import parse
import numpy as np


path = r'C:\Users\ma7054\Desktop\VIVO-23-09-chronic-scoresheet-20230313.xlsx'

#----OPEN FILES----
wb = load_workbook(filename = path)
xl = pd.ExcelFile(path)
list_name_group = [n for n in xl.sheet_names if 'Group' in n]
ws = wb[list_name_group[0]]



###-----FIND DATES-----
def search_date_in_string(value):
    match = re.search(r'\d{4}-\d{2}-\d{2}', value)
    if match == None:
        return
    else:
        date = datetime.strptime(match.group(), '%Y-%m-%d').date()
    return date
dates = []
for sheet_name in list_name_group:
    row_dates = []
    ws = wb[sheet_name]
    for cell in ws[2]:
        if cell.value != None:
            row_dates += [cell.value]
    ws_dates = [search_date_in_string(n) for n in row_dates if search_date_in_string(n) != None]
    if len(ws_dates)> len(dates):
        dates = [search_date_in_string(n) for n in row_dates if search_date_in_string(n) != None]
        dates_str = [date.strftime('%Y-%m-%d') for date in dates]

###------FIND VALUES OF EXPERIMENT------

def result_work_sheet(dates,dates_str,work_sheet_name):
    df = pd.read_excel(path,sheet_name=work_sheet_name,skiprows=2)
    df = df.dropna(subset=['Treatment','Cage Nb','Pyrat Nb'],how='all')
    weight = df.columns[df.columns.str.contains("Weight \(gr")].values
    score = df.columns[df.columns.str.contains("Total score")].values

    df_weight = df[weight].dropna(axis=1,how='all')


    df_score = df[score].iloc[:,0:len(df_weight.columns)]
    split = df_score.to_csv(header=None, index=False).strip('\n').split('\n')
    score = [n.rstrip().replace(",",", ") for n in split]


    #-----concat final dataframe------

    df_result = df.loc[:,['Treatment','Pyrat Nb','Cage Nb']]
    df_result['Group_info'] = work_sheet_name
    df_result['Score'] = score
    df_result['date_infection'] = dates[0]

    dates_good_format = ', '.join(dates_str)
    df_result['date_time'] = dates_good_format
    df_result = pd.concat([df_result,df_weight],axis=1)
    return df_result


list_df = []
for name in list_name_group:
    result = result_work_sheet(dates=dates,dates_str=dates_str,work_sheet_name=name)
    list_df+=[result]

final_result = pd.concat(list_df,axis=0)

final_result.to_excel(r'C:\Users\ma7054\Desktop\results.xlsx')